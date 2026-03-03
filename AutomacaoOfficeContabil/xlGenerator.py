from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


def gerarExcelSaida(grupos, nome_saida="Relatorio.xlsx"):
    # cria a planilha na memoria seleciona a planilha ativa e renomeia
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio Saida"

    # cria o cabeçalho da planilha
    ws.append(["Dia", "Série", "número", "valor contabil (R$)"])

    totalGeral = 0 # variavel da soma total de todos os dias

    # Percorre as chaves do dicionario recupera os devidos valores e adiciona a linha na planilha
    for (dia, serie) in sorted(grupos):
        dados = grupos[(dia, serie)]
        intervalo = f"{dados['min']}-{dados['max']}"
        valor = round(dados["total"], 2)

        totalGeral += valor
        ws.append([dia, serie, intervalo, valor])

    ws.append(["", "", "TOTAL", round(totalGeral, 2)]) # Adiciona o valor total do mes

    #Estilo da planilha
    headerFill = PatternFill("solid", fgColor="1F4E79")
    headerFont = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in ws[1]:
        c.fill = headerFill
        c.font = headerFont
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.border = border
            if c.column == 4:
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="center")

    ultimaLinha = ws.max_row
    for c in ws[ultimaLinha]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFF2CC")

    tabela = Table(displayName="TabelaRelatorio", ref=f"A1:D{ultimaLinha-1}")
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)

    wb.save(nome_saida)
    print(f"Planilha gerada com sucesso: {nome_saida}")

def gerarExcelServico(linhas, nome_saida="Relatorio.xlsx"):

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio Serviço"

    # cria Cabeçalho
    ws.append(["Dia", "número", "valor contabil (R$)"])

    totalGeral = 0

    # Ordena por Dia
    linhasOrdenadas = sorted(linhas, key=lambda x: x["Dia"])

    for item in linhasOrdenadas:
        dia = item["Dia"]
        numero = item["número"]
        valor = round(item["valor contabil (R$)"], 2)

        totalGeral += valor
        ws.append([dia, numero, valor])

    # Linha TOTAL
    ws.append(["", "TOTAL", round(totalGeral, 2)])

    # Estilo da planilha
    headerFill = PatternFill("solid", fgColor="1F4E79")
    headerFont = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Estilo cabeçalho
    for c in ws[1]:
        c.fill = headerFill
        c.font = headerFont
        c.alignment = Alignment(horizontal="center")
        c.border = border

    # Estilo corpo
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.border = border
            if c.column == 3:
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="center")

    # Destacar linha TOTAL
    ultimaLinha = ws.max_row
    for c in ws[ultimaLinha]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFF2CC")

    # Criar tabela formatada
    tabela = Table(displayName="TabelaRelatorioServico",
                   ref=f"A1:C{ultimaLinha-1}")

    estilo = TableStyleInfo(name="TableStyleMedium9",
                            showRowStripes=True)

    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)

    wb.save(nome_saida)
    print(f"Planilha gerada com sucesso: {nome_saida}")
